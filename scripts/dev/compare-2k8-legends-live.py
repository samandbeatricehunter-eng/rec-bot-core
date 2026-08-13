"""Compare All-Pro Football 2K8 spreadsheet vs live rec_legend_catalog."""
from __future__ import annotations

import collections
import json
import re
import urllib.request
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\josh_\Downloads\all-pro_football_2k8.xlsx")
CATALOG_JSON = Path(__file__).with_name("legend-catalog-live.json")


def normalize(name: str) -> str:
    n = name.lower().strip().replace("’", "'").replace("‘", "'")
    n = n.replace("ñ", "n").replace("Ñ", "n")
    n = re.sub(r'\s*"[^"]+"\s*', " ", n)
    n = re.sub(r"\s+", " ", n).replace(".", "").strip()
    return n


def first_last(name: str) -> str:
    parts = normalize(name).split()
    if len(parts) >= 2:
        return f"{parts[0]} {parts[-1]}"
    return normalize(name)


def parse_sheet():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["All-Pro Football 2K8"]
    players = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        players.append(
            {
                "pos": str(row[0] or "").strip(),
                "name": str(row[1]).strip(),
                "class": str(row[2] or "").strip(),
                "abilities": str(row[3] or "").strip(),
            }
        )
    skill_rows = wb["Skill Chart"].max_row - 1
    return players, skill_rows


def match_against(players, cat_rows, label: str):
    by_norm = {normalize(c["name"]): c for c in cat_rows}
    by_fl: dict[str, list] = collections.defaultdict(list)
    for c in cat_rows:
        by_fl[first_last(c["name"])].append(c)

    exact, fuzzy, missing = [], [], []
    for p in players:
        k = normalize(p["name"])
        fl = first_last(p["name"])
        if k in by_norm:
            exact.append((p, by_norm[k]))
            continue
        if fl in by_fl:
            fuzzy.append((p, by_fl[fl][0], "first_last"))
            continue
        # e.g. sheet "Joe Greene" vs catalog "Mean Joe Greene"
        hit = None
        plast = normalize(p["name"]).split()[-1]
        pfirst = normalize(p["name"]).split()[0]
        for cn, c in by_norm.items():
            cparts = cn.split()
            if not cparts or cparts[-1] != plast:
                continue
            if pfirst in cparts or cparts[0] == pfirst:
                hit = c
                break
        if hit:
            fuzzy.append((p, hit, "partial"))
            continue
        missing.append(p)

    matched = [(p, c) for p, c in exact] + [(p, c) for p, c, _ in fuzzy]
    print(f"\n=== vs {label} ({len(cat_rows)} rows / {len({normalize(c['name']) for c in cat_rows})} distinct names) ===")
    print(f"exact={len(exact)} fuzzy={len(fuzzy)} IN_CATALOG={len(exact)+len(fuzzy)} missing={len(missing)}")
    print("matched_by_tier", dict(collections.Counter(p["class"] for p, _ in matched)))
    print("missing_by_tier", dict(collections.Counter(p["class"] for p in missing)))
    print("MATCHED:")
    for p, c in exact:
        print(f"  [{p['class']}] {p['name']} ({p['pos']}) == {c['name']} [{c['game_scope']}/{c['position']}]")
    for p, c, how in fuzzy:
        print(f"  [{p['class']}] {p['name']} ({p['pos']}) ~= {c['name']} [{c['game_scope']}/{c['position']}] ({how})")
    return exact, fuzzy, missing


def main():
    players, skill_rows = parse_sheet()
    catalog = json.loads(CATALOG_JSON.read_text(encoding="utf-8"))

    print(f"SPREADSHEET_TOTAL={len(players)}")
    print("BY_CLASS", dict(collections.Counter(p["class"] for p in players)))
    print("BY_POS", dict(sorted(collections.Counter(p["pos"] for p in players).items())))
    print(f"SKILL_CHART_ABILITIES={skill_rows}")
    print(f"CATALOG_ROWS={len(catalog)}")
    print("CATALOG_BY_SCOPE", dict(collections.Counter(c["game_scope"] for c in catalog)))

    madden = [c for c in catalog if c["game_scope"] == "madden"]
    match_against(players, madden, "madden only")
    _, _, missing_any = match_against(players, catalog, "any scope (madden+cfb)")

    # Gold missing from any scope — highest-value gaps
    gold_missing = [p for p in missing_any if p["class"] == "Gold"]
    print(f"\nGOLD NOT IN ANY CATALOG ({len(gold_missing)}):")
    for p in gold_missing:
        print(f"  {p['name']} ({p['pos']})")


if __name__ == "__main__":
    main()
